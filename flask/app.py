from flask import Flask, request, jsonify
from flask_cors import CORS
import psycopg2
import psycopg2.extras
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler
import threading
import time

app = Flask(__name__)
CORS(app)

def get_db():
    return psycopg2.connect(
        host='localhost', database='hydroponic_db',
        user='hydro_user', password='hydro123', port=5432
    )

_pulse_lock = {
    'pompa_2': False,
    'pompa_3': False,
}

def auto_off_pompa(nama_pompa, durasi_detik, keterangan=''):
    def _off():
        time.sleep(durasi_detik)
        try:
            conn = get_db()
            cur  = conn.cursor()
            cur.execute("UPDATE pompa_status SET status='off', updated_at=NOW() WHERE nama_pompa=%s", (nama_pompa,))
            cur.execute("INSERT INTO log_aktivitas (pompa, aksi, keterangan, trigger_type) VALUES (%s, 'off', %s, 'auto')", (nama_pompa, keterangan or f'Auto-off setelah {durasi_detik} detik'))
            conn.commit()
            cur.close(); conn.close()
            if nama_pompa in _pulse_lock:
                _pulse_lock[nama_pompa] = False
        except Exception as e:
            print(f"Error auto-off {nama_pompa}: {e}")
            if nama_pompa in _pulse_lock:
                _pulse_lock[nama_pompa] = False
    threading.Thread(target=_off, daemon=True).start()

def pulse_dosing(nama_pompa, fungsi, ph_sekarang, target, durasi=5, jeda=60):
    if _pulse_lock.get(nama_pompa):
        print(f"[PULSE] {nama_pompa} masih dalam siklus, skip")
        return False
    _pulse_lock[nama_pompa] = True
    def _pulse():
        try:
            conn = get_db()
            cur  = conn.cursor()
            cur.execute("UPDATE pompa_status SET status='on', updated_at=NOW() WHERE nama_pompa=%s", (nama_pompa,))
            cur.execute("INSERT INTO log_aktivitas (pompa, aksi, keterangan, trigger_type, kotak) VALUES (%s, 'on', %s, 'auto', 'kotak_1')", (nama_pompa, f'{fungsi}: pH {ph_sekarang} → target {target} | pulse {durasi} detik'))
            conn.commit()
            cur.close(); conn.close()
            print(f"[PULSE] {nama_pompa} ON - {durasi} detik")
            time.sleep(durasi)
            conn = get_db()
            cur  = conn.cursor()
            cur.execute("UPDATE pompa_status SET status='off', updated_at=NOW() WHERE nama_pompa=%s", (nama_pompa,))
            cur.execute("INSERT INTO log_aktivitas (pompa, aksi, keterangan, trigger_type) VALUES (%s, 'off', %s, 'auto')", (nama_pompa, f'Auto-off setelah pulse {durasi} detik'))
            conn.commit()
            cur.close(); conn.close()
            print(f"[PULSE] {nama_pompa} OFF - tunggu {jeda} detik mixing")
            time.sleep(jeda)
            _pulse_lock[nama_pompa] = False
            print(f"[PULSE] {nama_pompa} unlock - siap cek ulang")
        except Exception as e:
            print(f"Error pulse {nama_pompa}: {e}")
            _pulse_lock[nama_pompa] = False
    threading.Thread(target=_pulse, daemon=True).start()
    return True

def isi_air_mixing():
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT status FROM pompa_status WHERE nama_pompa='pompa_1'")
        row = cur.fetchone()
        if row and row[0] == 'on':
            print("[ISI AIR] P1 sudah ON, skip")
            cur.close(); conn.close()
            return
        cur.execute("UPDATE pompa_status SET status='on', updated_at=NOW() WHERE nama_pompa='pompa_1'")
        cur.execute("INSERT INTO log_aktivitas (pompa, aksi, keterangan, trigger_type, kotak) VALUES ('pompa_1', 'on', 'Water level kosong - isi air', 'auto', 'kotak_1')")
        conn.commit()
        cur.close(); conn.close()
        print("[ISI AIR] P1 ON - tunggu WL terdeteksi")
    except Exception as e:
        print(f"Error isi air: {e}")

def matikan_p1():
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT status FROM pompa_status WHERE nama_pompa='pompa_1'")
        row = cur.fetchone()
        if row and row[0] == 'on':
            cur.execute("UPDATE pompa_status SET status='off', updated_at=NOW() WHERE nama_pompa='pompa_1'")
            cur.execute("INSERT INTO log_aktivitas (pompa, aksi, keterangan, trigger_type, kotak) VALUES ('pompa_1', 'off', 'Water level terdeteksi - P1 OFF', 'auto', 'kotak_1')")
            conn.commit()
            print("[ISI AIR] P1 OFF - WL terdeteksi")
        cur.close(); conn.close()
    except Exception as e:
        print(f"Error matikan P1: {e}")

def kontrol_ph_mixing(ph, ph_min, ph_max, durasi=5, jeda=60):
    if ph < ph_min:
        print(f"[pH] {ph} < {ph_min} → pulse P2 (pH Up)")
        pulse_dosing('pompa_2', 'pH Up', ph, ph_min, durasi, jeda)
    elif ph > ph_max:
        print(f"[pH] {ph} > {ph_max} → pulse P3 (pH Down)")
        pulse_dosing('pompa_3', 'pH Down', ph, ph_max, durasi, jeda)
    else:
        try:
            conn = get_db()
            cur  = conn.cursor()
            cur.execute("UPDATE pompa_status SET status='off', updated_at=NOW() WHERE nama_pompa IN ('pompa_2','pompa_3') AND status='on'")
            conn.commit()
            cur.close(); conn.close()
            print(f"[pH] {ph} normal → P2 & P3 OFF")
        except Exception as e:
            print(f"Error kontrol pH: {e}")

def kontrol_distribusi(kotak, pompa, ph, tds, ph_min, ph_max, tds_min, tds_max):
    butuh = (ph < ph_min or ph > ph_max or tds < tds_min)
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT status FROM pompa_status WHERE nama_pompa=%s", (pompa,))
        row = cur.fetchone()
        status_sekarang = row[0] if row else 'off'
        if butuh and status_sekarang == 'off':
            cur.execute("UPDATE pompa_status SET status='on', updated_at=NOW() WHERE nama_pompa=%s", (pompa,))
            cur.execute("INSERT INTO log_aktivitas (pompa, aksi, keterangan, trigger_type, kotak) VALUES (%s, 'on', %s, 'auto', %s)", (pompa, f'pH {ph} / TDS {tds} kurang - distribusi dari mixing', kotak))
            print(f"[DISTRIBUSI] {pompa} ON → {kotak}")
        elif not butuh and status_sekarang == 'on':
            cur.execute("UPDATE pompa_status SET status='off', updated_at=NOW() WHERE nama_pompa=%s", (pompa,))
            cur.execute("INSERT INTO log_aktivitas (pompa, aksi, keterangan, trigger_type, kotak) VALUES (%s, 'off', 'pH & TDS normal - pompa OFF', 'auto', %s)", (pompa, kotak))
            print(f"[DISTRIBUSI] {pompa} OFF → {kotak} normal")
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print(f"Error distribusi {kotak}: {e}")

def cek_mixing_tank():
    print(f"[{datetime.now()}] Scheduler: Cek Mixing Tank...")
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT mode FROM mode_sistem ORDER BY id DESC LIMIT 1")
        mode_row = cur.fetchone()
        mode = mode_row[0] if mode_row else 'semi'
        if mode == 'manual':
            cur.close(); conn.close(); return
        cur.execute("SELECT ph, tds FROM sensor_data WHERE kotak='kotak_1' ORDER BY timestamp DESC LIMIT 1")
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close(); return
        ph, tds = row
        cur.execute("SELECT * FROM pengaturan ORDER BY id DESC LIMIT 1")
        s = cur.fetchone()
        cur.close(); conn.close()
        if not s: return
        ph_min  = s[1]; ph_max = s[2]
        tds_min = s[3]; tds_max = s[4]
        durasi  = s[8] if len(s) > 8 and s[8] else 5
        jeda    = s[9] if len(s) > 9 and s[9] else 60

        kontrol_ph_mixing(ph, ph_min, ph_max, durasi, jeda)
    except Exception as e:
        print(f"Error scheduler mixing: {e}")

def cek_kotak_tanam():
    print(f"[{datetime.now()}] Scheduler: Cek Kotak Tanam...")
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT mode FROM mode_sistem ORDER BY id DESC LIMIT 1")
        mode_row = cur.fetchone()
        mode = mode_row[0] if mode_row else 'semi'
        if mode == 'manual':
            cur.close(); conn.close(); return
        cur.execute("SELECT * FROM pengaturan ORDER BY id DESC LIMIT 1")
        s = cur.fetchone()
        cur.close(); conn.close()
        if not s: return
        ph_min  = s[1]; ph_max = s[2]
        tds_min = s[3]; tds_max = s[4]
        kotak_pompa = {'kotak_2': 'pompa_4', 'kotak_3': 'pompa_5', 'kotak_4': 'pompa_6'}
        for kotak, pompa in kotak_pompa.items():
            conn = get_db()
            cur  = conn.cursor()
            cur.execute("SELECT ph, tds FROM sensor_data WHERE kotak=%s ORDER BY timestamp DESC LIMIT 1", (kotak,))
            row = cur.fetchone()
            cur.close(); conn.close()
            if not row: continue
            ph, tds = row
            kontrol_distribusi(kotak, pompa, ph, tds, ph_min, ph_max, tds_min, tds_max)
    except Exception as e:
        print(f"Error scheduler kotak: {e}")

scheduler = BackgroundScheduler()
def get_interval():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT interval_mixing, interval_kotak FROM pengaturan ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close(); conn.close()
        return (row[0] or 360, row[1] or 180) if row else (360, 180)
    except: return (360, 180)
_im, _ik = get_interval()
scheduler.add_job(cek_mixing_tank, trigger='interval', minutes=_im, id='cek_mixing', replace_existing=True)
scheduler.add_job(cek_kotak_tanam, trigger='interval', minutes=_ik, id='cek_kotak',  replace_existing=True)
scheduler.start()

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'time': datetime.now().isoformat()}), 200

@app.route('/api/mode', methods=['GET'])
def get_mode():
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM mode_sistem ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close(); conn.close()
        return jsonify(dict(row) if row else {'mode': 'semi'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mode', methods=['POST'])
def set_mode():
    data = request.get_json()
    mode = data.get('mode')
    if mode not in ('auto', 'semi', 'manual'):
        return jsonify({'error': 'Mode harus auto, semi, atau manual'}), 400
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("INSERT INTO mode_sistem (mode) VALUES (%s)", (mode,))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({'status': 'ok', 'mode': mode}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sensor', methods=['POST'])
def post_sensor():
    data        = request.get_json()
    kotak       = data.get('kotak')
    ph          = data.get('ph')
    tds         = data.get('tds')
    water_level = data.get('water_level', False)
    if not kotak or ph is None or tds is None:
        return jsonify({'error': 'Field kotak, ph, tds wajib diisi'}), 400
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("INSERT INTO sensor_data (kotak, ph, tds, water_level) VALUES (%s,%s,%s,%s)", (kotak, ph, tds, water_level))
        conn.commit()
        cur.execute("SELECT mode FROM mode_sistem ORDER BY id DESC LIMIT 1")
        mode_row = cur.fetchone()
        mode = mode_row[0] if mode_row else 'semi'
        cur.execute("SELECT * FROM pengaturan ORDER BY id DESC LIMIT 1")
        s = cur.fetchone()
        cur.close(); conn.close()
        aksi_pompa = []
        if mode == 'auto' and s:
            ph_min  = s[1]; ph_max = s[2]
            tds_min = s[3]; tds_max = s[4]
            durasi  = s[8] if len(s) > 8 and s[8] else 5
            jeda    = s[9] if len(s) > 9 and s[9] else 60
            if kotak == 'kotak_1':

                    if ph < ph_min:
                        ok = pulse_dosing('pompa_2', 'pH Up', ph, ph_min, durasi, jeda)
                        if ok: aksi_pompa.append(f'pompa_2 PULSE {durasi}s - pH {ph} rendah')
                    elif ph > ph_max:
                        ok = pulse_dosing('pompa_3', 'pH Down', ph, ph_max, durasi, jeda)
                        if ok: aksi_pompa.append(f'pompa_3 PULSE {durasi}s - pH {ph} tinggi')
                    else:
                        aksi_pompa.append(f'pH {ph} normal')
            elif kotak in ('kotak_2', 'kotak_3', 'kotak_4'):
                pompa_map = {'kotak_2': 'pompa_4', 'kotak_3': 'pompa_5', 'kotak_4': 'pompa_6'}
                pompa = pompa_map[kotak]
                kontrol_distribusi(kotak, pompa, ph, tds, ph_min, ph_max, tds_min, tds_max)
                butuh = (ph < ph_min or ph > ph_max or tds < tds_min)
                aksi_pompa.append(f'{pompa} {"ON" if butuh else "OFF"}')
        return jsonify({'status': 'ok', 'kotak': kotak, 'mode': mode, 'otomasi': aksi_pompa}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sensor/latest', methods=['GET'])
def get_sensor_latest():
    kotak_list = ['kotak_1', 'kotak_2', 'kotak_3', 'kotak_4']
    result = {}
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        for kotak in kotak_list:
            cur.execute("SELECT kotak,ph,tds,water_level,timestamp FROM sensor_data WHERE kotak=%s ORDER BY timestamp DESC LIMIT 1", (kotak,))
            row = cur.fetchone()
            if row:
                result[kotak] = {
                    'kotak'      : row['kotak'],
                    'ph'         : float(row['ph']),
                    'tds'        : float(row['tds']),
                    'water_level': row['water_level'],
                    'timestamp'  : row['timestamp'].isoformat()
                }
        cur.close(); conn.close()
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sensor/history', methods=['GET'])
def get_sensor_history():
    kotak = request.args.get('kotak', 'kotak_1')
    days  = int(request.args.get('days', 7))
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT DATE(timestamp) as tanggal,
                   AVG(ph) as ph_avg, AVG(tds) as tds_avg,
                   MIN(ph) as ph_min, MAX(ph) as ph_max,
                   MIN(tds) as tds_min, MAX(tds) as tds_max
            FROM sensor_data
            WHERE kotak=%s AND timestamp >= NOW() - INTERVAL '%s days'
            GROUP BY DATE(timestamp)
            ORDER BY tanggal ASC
        """, (kotak, days))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{
            'tanggal': str(r['tanggal']),
            'ph_avg' : round(float(r['ph_avg']), 2),
            'tds_avg': round(float(r['tds_avg']), 2),
            'ph_min' : round(float(r['ph_min']), 2),
            'ph_max' : round(float(r['ph_max']), 2),
            'tds_min': round(float(r['tds_min']), 2),
            'tds_max': round(float(r['tds_max']), 2),
        } for r in rows]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sensor/weekly', methods=['GET'])
def get_sensor_weekly():
    kotak = request.args.get('kotak', 'kotak_1')
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT AVG(ph) as ph_avg, AVG(tds) as tds_avg,
                   MIN(ph) as ph_min, MAX(ph) as ph_max,
                   MIN(tds) as tds_min, MAX(tds) as tds_max,
                   COUNT(*) as total_data
            FROM sensor_data
            WHERE kotak=%s AND timestamp >= NOW() - INTERVAL '7 days'
            AND ph >= 0 AND ph <= 14 AND tds > 0
        """, (kotak,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row and row['total_data'] > 0:
            return jsonify({
                'kotak'     : kotak,
                'ph_avg'    : round(float(row['ph_avg']), 2),
                'tds_avg'   : round(float(row['tds_avg']), 2),
                'ph_min'    : round(float(row['ph_min']), 2),
                'ph_max'    : round(float(row['ph_max']), 2),
                'tds_min'   : round(float(row['tds_min']), 2),
                'tds_max'   : round(float(row['tds_max']), 2),
                'total_data': row['total_data']
            }), 200
        return jsonify({'kotak': kotak, 'message': 'Belum ada data minggu ini'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pompa/status', methods=['GET'])
def get_pompa_status():
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM pompa_status ORDER BY id")
        rows = cur.fetchall()
        cur.close(); conn.close()
        result = {}
        for r in rows:
            result[r['nama_pompa']] = {
                'fungsi'    : r['fungsi'],
                'status'    : r['status'],
                'kotak'     : r.get('kotak'),
                'updated_at': r['updated_at'].isoformat()
            }
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pompa/control', methods=['POST'])
def kontrol_pompa():
    data         = request.get_json()
    pompa        = data.get('pompa')
    status       = data.get('status')
    trigger_type = data.get('trigger', 'manual')
    if not pompa or status not in ('on', 'off'):
        return jsonify({'error': 'Field pompa dan status wajib diisi'}), 400
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT mode FROM mode_sistem ORDER BY id DESC LIMIT 1")
        mode_row = cur.fetchone()
        mode = mode_row[0] if mode_row else 'semi'
        if mode == 'auto' and trigger_type == 'manual':
            return jsonify({'error': 'Mode AUTO aktif, manual dikunci!'}), 403
        if status == 'on':
            if pompa == 'pompa_2':
                cur.execute("SELECT status FROM pompa_status WHERE nama_pompa='pompa_3'")
                p3 = cur.fetchone()
                if p3 and p3[0] == 'on':
                    return jsonify({'error': 'Pompa 3 aktif, pompa 2 dikunci!'}), 403
            if pompa == 'pompa_3':
                cur.execute("SELECT status FROM pompa_status WHERE nama_pompa='pompa_2'")
                p2 = cur.fetchone()
                if p2 and p2[0] == 'on':
                    return jsonify({'error': 'Pompa 2 aktif, pompa 3 dikunci!'}), 403
        cur.execute("UPDATE pompa_status SET status=%s, updated_at=NOW() WHERE nama_pompa=%s", (status, pompa))
        cur.execute("INSERT INTO log_aktivitas (pompa, aksi, keterangan, trigger_type) VALUES (%s,%s,%s,%s)", (pompa, status, f'Kontrol {trigger_type}', trigger_type))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({'status': 'ok', 'pompa': pompa, 'aksi': status}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pengaturan', methods=['GET'])
def get_pengaturan():
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM pengaturan ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close(); conn.close()
        return jsonify(dict(row) if row else {}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pengaturan', methods=['POST'])
def post_pengaturan():
    data = request.get_json()
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("""
            UPDATE pengaturan SET
                ph_min=%s, ph_max=%s, tds_min=%s, tds_max=%s,
                interval_mixing=%s, interval_kotak=%s,
                faktor_ph=%s, faktor_tds=%s
            WHERE id=(SELECT id FROM pengaturan ORDER BY id DESC LIMIT 1)
        """, (
            data.get('ph_min',    5.5),
            data.get('ph_max',    6.5),
            data.get('tds_min',   800),
            data.get('tds_max',   1200),
            data.get('interval_mixing', 360),
            data.get('interval_kotak',  180),
            data.get('faktor_ph',  5.0),
            data.get('faktor_tds', 60.0),
        ))
        conn.commit()
        cur.close(); conn.close()
        scheduler.reschedule_job('cek_mixing', trigger='interval', minutes=data.get('interval_mixing', 360))
        scheduler.reschedule_job('cek_kotak',  trigger='interval', minutes=data.get('interval_kotak',  180))
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/jadwal', methods=['GET'])
def get_jadwal():
    try:
        jobs = []
        for job in scheduler.get_jobs():
            jobs.append({'id': job.id, 'next_run': job.next_run_time.isoformat() if job.next_run_time else None})
        return jsonify({'jobs': jobs}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/log', methods=['GET'])
def get_log():
    limit        = int(request.args.get('limit', 20))
    trigger_type = request.args.get('trigger')
    kotak        = request.args.get('kotak')
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        query  = "SELECT * FROM log_aktivitas WHERE 1=1"
        params = []
        if trigger_type:
            query += " AND trigger_type=%s"
            params.append(trigger_type)
        if kotak:
            query += " AND kotak=%s"
            params.append(kotak)
        query += " ORDER BY timestamp DESC LIMIT %s"
        params.append(limit)
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{
            'id'          : r['id'],
            'pompa'       : r['pompa'],
            'aksi'        : r['aksi'],
            'keterangan'  : r['keterangan'],
            'trigger_type': r.get('trigger_type', 'auto'),
            'kotak'       : r.get('kotak'),
            'timestamp'   : r['timestamp'].isoformat()
        } for r in rows]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '')
    password = data.get('password', '')
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT id, username, role FROM users WHERE username=%s AND password=%s", (username, password))
        user = cur.fetchone()
        cur.close(); conn.close()
        if user:
            return jsonify({'status': 'ok', 'id': user[0], 'username': user[1], 'role': user[2]}), 200
        else:
            return jsonify({'status': 'error', 'message': 'Username atau password salah'}), 401
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/me', methods=['POST'])
def me():
    data = request.get_json()
    username = data.get('username', '')
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT id, username, role FROM users WHERE username=%s", (username,))
        user = cur.fetchone()
        cur.close(); conn.close()
        if user:
            return jsonify({'id': user[0], 'username': user[1], 'role': user[2]}), 200
        else:
            return jsonify({'status': 'error'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


import csv
import io
from flask import Response

@app.route('/api/sensor/export', methods=['GET'])
def export_sensor_csv():
    kotak  = request.args.get('kotak', 'kotak_1')
    dari   = request.args.get('dari')
    sampai = request.args.get('sampai')
    try:
        conn = get_db()
        cur  = conn.cursor()
        query = "SELECT timestamp, ph, tds, water_level FROM sensor_data WHERE kotak = %s"
        params = [kotak]
        if dari:
            query += " AND DATE(timestamp) >= %s"
            params.append(dari)
        if sampai:
            query += " AND DATE(timestamp) <= %s"
            params.append(sampai)
        query += " ORDER BY timestamp ASC"
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close(); conn.close()
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Sensor"
        ws.append(['Timestamp', 'pH', 'TDS', 'Water Level'])
        for row in rows:
            ws.append([str(row[0]).replace('T', ' ').split('.')[0], float(row[1]), float(row[2]), str(row[3])])
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"sensor_{kotak}_{dari}_sampai_{sampai}.xlsx"
        return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={filename}', 'Access-Control-Allow-Origin': '*'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=False)
